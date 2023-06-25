import re
from mystr import mystr
from openpyxl import load_workbook

class my_excel:
    def __init__(self,path,sh):
        self.path=path  #excel文件路径
        self.sh=sh     #excel  sheet

    #获取EXCEL数据
    def get_date_toDict(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.path)
        sh1=wb[self.sh]
        content= {}
        title={}
        for i,row  in  enumerate(sh1.rows):
            if i==0 :
                for j,cell in enumerate(row):
                    # content[cell.value]=[cell.value]
                    content[cell.value] = []
                    title[j]=cell.value
            else:
                for j,cell in enumerate(row):
                    content[title[j]].append(cell.value)
        return title, content

    def sheel_upper(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.path)
        sh1 = wb[self.sh]
        for i, row in enumerate(sh1.rows):
            for j, cell in enumerate(row):
                if cell.value != None:
                    cell.value=cell.value.upper()
        # wb.save(self.sh+'(修改).xlsx')
        wb.save(f'../create_data/{self.sh}(修改).xlsx')


    def get_date_toStr(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.path)
        sh1=wb[self.sh]
        content=""
        title=""
        for i,row  in  enumerate(sh1.rows):
            if i==0 :
                for j,cell in enumerate(row):
                    # content[cell.value]=[cell.value]
                    if cell.value != None:
                        title += str(cell.value)+'\n'
            else:
                for j,cell in enumerate(row):
                    if cell.value !=None:
                        content+=str(cell.value)+'\n'
        return title, content

    #取出oracle包中的procedure  返回[]
    def get_procedures(self,str1:str):
        arrs=[]
        arr=None
        str_rows=str1.split('\n')
        for i in range(len(str_rows)):
            tmp=str_rows[i].strip().upper()
            # print('tmp:',tmp)
            # print("tmp[0:9]:", tmp[0:9])
            if tmp[0:9] =='PROCEDURE':
                if  arr !=None:
                    arrs.append(arr)
                    arr = str_rows[i].strip().upper()+'\n'
                else :
                    arr=str_rows[i].strip().upper()
            else :
                if  arr !=None:
                    arr+= str_rows[i].strip()+'\n'
                else :
                    pass
        return arrs

    #EXCEL列数据拆分   将sheet列  按字符串split_str 拆分至 sheets 多列中
    def sheet_split(self,sheet,split_str,*sheets):
        wb = load_workbook(self.path)
        sh1 = wb[self.sh]
        print(self.path,self.sh,sh1)
        content = []
        title =  []
        for i, row in enumerate(sh1.rows):
            if i == 0:
                for j, cell in enumerate(row):
                    # content[cell.value]=[cell.value]
                    if cell.value != None:
                        title += str(cell.value) + '\n'
            else:
                for j, cell in enumerate(row):
                    if cell.value != None:
                        content += str(cell.value) + '\n'
        return title, content

if __name__ == '__main__':
    # my_excel1=my_excel('../base_data/ora_pac.xlsx', 'PAG_XWH_WG_MON')
    # title, content=my_excel1.get_date_toStr()
    # print(title,content)
    # my_excel1.get_procedures(content)
    my_excel1=my_excel('../base_data/ora_pac.xlsx', '字典')
    # my_excel1.sheet_split("原始数据",".","O_USER","O_TB")
    my_excel1.sheel_upper()
